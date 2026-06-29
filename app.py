import streamlit as st
import pandas as pd
import re
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==========================================
# CONFIGURACIÓN DE LA PÁGINA
# ==========================================
st.set_page_config(page_title="Conciliador de Logística", layout="wide")
st.title("📊 Conciliador Automático de Logística")
st.write("Sube el archivo de descarga para realizar la conciliación automática.")

# ==========================================
# FUNCIONES DE LIMPIEZA
# ==========================================
def extraer_ean(txt):
    s = str(txt).strip()
    if s.endswith('.0'): s = s[:-2]
    match = re.search(r'\[(\d+)\]', s)
    if match: return match.group(1).lstrip('0') or '0'
    match_num = re.search(r'(\d+)', s)
    if match_num: return match_num.group(1).lstrip('0') or '0'
    res_s = s.lstrip('0')
    return res_s if res_s else '0'

def extraer_pedido_glosa(txt):
    if pd.isna(txt): return "SIN_PEDIDO"
    texto = str(txt).upper().strip()
    if texto.endswith('.0'): texto = texto[:-2]
    match_p = re.search(r'(P\d+)', texto)
    if match_p: return match_p.group(1)
    match_oc = re.search(r'(?:OC|ORDEN|PEDIDO)?\s*(\d{4,})', texto)
    if match_oc: return f"P{match_oc.group(1)}"
    return "SIN_PEDIDO"

def preservar_comprobante(txt):
    if pd.isna(txt) or str(txt).strip() in ['', 'nan', 'None', 'NAN']: return "SIN_DOC"
    texto = str(txt).strip().upper()
    if texto.endswith(".0"): texto = texto[:-2]
    return texto

def normalizar_comprobante_dinamico(txt, pedido_ref=None, es_sdp=False, dict_prov=None, dict_pref=None):
    texto = str(txt).strip().upper()
    if texto in ['NAN', '', 'NONE', 'NAN ', 'SIN_DOC']: return None
    if "CON" in texto:
        numeros_con = "".join(re.findall(r'\d+', texto))
        return f"CON-{numeros_con}"
    
    if texto.endswith(".0"): texto = texto[:-2]
    numeros = "".join(re.findall(r'\d+', texto))
    if not numeros: return None
    
    correlativo = numeros.zfill(8)[-8:]
    prefijo_base = "F039" 
    
    if pedido_ref and pedido_ref != "SIN_PEDIDO" and dict_prov and dict_pref:
        nom_proveedor = dict_prov.get(str(pedido_ref).strip())
        if nom_proveedor and nom_proveedor in dict_pref:
            prefijo_base = dict_pref[nom_proveedor]
        
    if es_sdp or "FC39" in texto or "RRE" in texto or "CREDITO" in texto or "DEV" in texto:
        if len(prefijo_base) >= 2:
            prefijo_nc = prefijo_base[0] + "C" + prefijo_base[2:]
        else:
            prefijo_nc = "FC39"
        return f"{prefijo_nc}-{correlativo}"
        
    return f"{prefijo_base}-{correlativo}"

# ==========================================
# INTERFAZ DE CARGA
# ==========================================
uploaded_file = st.file_uploader("📂 Selecciona tu archivo Excel", type=["xlsx"])

if uploaded_file is not None:
    if st.button("🚀 Iniciar Conciliación"):
        with st.spinner('Procesando la conciliación...'):
            try:
                # PASO 1: CARGA DE ARCHIVOS EN MEMORIA
                xls = pd.ExcelFile(uploaded_file)
                lista_hojas = xls.sheet_names
                hoja_proveedores = next((h for h in lista_hojas if "PROV" in h.strip().upper()), None)
                
                if not hoja_proveedores:
                    st.error(f"No se encontró la pestaña 'PROVEEDORES'. Hojas disponibles: {lista_hojas}")
                    st.stop()
                if "DETALLE_FACTURAS" not in lista_hojas:
                    st.error("No se encontró la pestaña obligatoria 'DETALLE_FACTURAS'.")
                    st.stop()

                df_conta_orig = pd.read_excel(xls, sheet_name="CONTABILIDAD")
                df_kardex_orig = pd.read_excel(xls, sheet_name="LOGISTICA")
                df_proveedores_orig = pd.read_excel(xls, sheet_name=hoja_proveedores)
                df_detalle_orig = pd.read_excel(xls, sheet_name="DETALLE_FACTURAS")

                # Diccionarios de apoyo
                df_proveedores_orig.columns = [str(c).strip().upper() for c in df_proveedores_orig.columns]
                dict_prefijos = {}
                for _, row in df_proveedores_orig.dropna(subset=['REFERENCIA', 'PREFIJO']).iterrows():
                    dict_prefijos[str(row['REFERENCIA']).strip().upper()] = str(row['PREFIJO']).strip().upper()
                
                df_kardex_validos = df_kardex_orig[df_kardex_orig['CodTGu'].isin(['IC', 'SDP'])]
                registros_ped_prov = []
                for _, row in df_kardex_validos.dropna(subset=['PO', 'REFERENCIA']).iterrows():
                    ped_k = str(row['PO']).strip()
                    prov_k = str(row['REFERENCIA']).strip().upper()
                    if [ped_k, prov_k] not in registros_ped_prov and ped_k != '':
                        registros_ped_prov.append([ped_k, prov_k])
                dict_pedido_a_proveedor = dict(registros_ped_prov)

                # PASO 3: PIPELINE DE DETALLE_FACTURAS
                df_detalle = df_detalle_orig.copy()
                df_detalle['Nro. Documento'] = df_detalle['Nro. Documento'].ffill()
                df_detalle['Glosa'] = df_detalle['Glosa'].ffill()
                if 'Nombre de la empresa a mostrar en la factura' in df_detalle.columns:
                    df_detalle['Nombre de la empresa a mostrar en la factura'] = df_detalle['Nombre de la empresa a mostrar en la factura'].ffill()
                
                df_detalle = df_detalle.dropna(subset=['Líneas de factura/Producto/Tipo de producto'])
                df_detalle = df_detalle[df_detalle['Líneas de factura/Producto/Tipo de producto'].astype(str).str.strip() != '']
                
                df_detalle['PEDIDO_CORRECTO'] = df_detalle['Glosa'].apply(extraer_pedido_glosa)
                df_detalle['EAN_LIMPIO'] = df_detalle['Líneas de factura/Producto'].apply(extraer_ean)
                df_detalle['CANT_ABS'] = df_detalle['Líneas de factura/Cantidad'].apply(lambda x: round(abs(float(str(x).replace(',', '.'))), 4) if pd.notnull(x) else 0.0)
                df_detalle['PROV_LIMPIO'] = df_detalle['Nombre de la empresa a mostrar en la factura'].apply(lambda x: str(x).strip().upper())
                
                costos_calculados = []
                for _, row in df_detalle.iterrows():
                    try:
                        total_fact = float(str(row['Líneas de factura/Total']).replace(',', '.'))
                        cantidad_fact = float(str(row['Líneas de factura/Cantidad']).replace(',', '.'))
                        costo_u_sin_igv = (total_fact / cantidad_fact) / 1.18 if cantidad_fact != 0 else 0.0
                    except:
                        costo_u_sin_igv = 0.0
                    costos_calculados.append(costo_u_sin_igv)
                df_detalle['COSTO_UNITARIO_NETO'] = costos_calculados
                df_detalle['FACTURA_NORMALIZADA'] = df_detalle['Nro. Documento'].apply(preservar_comprobante)

                # PASO 4: ÍNDICES COMPARTIDOS
                lineas_detalle = []
                for _, row in df_detalle.iterrows():
                    lineas_detalle.append({
                        'pedido': str(row['PEDIDO_CORRECTO']).strip(),
                        'ean': str(row['EAN_LIMPIO']).strip(),
                        'proveedor': str(row['PROV_LIMPIO']).strip(),
                        'cant': row['CANT_ABS'],
                        'costo': row['COSTO_UNITARIO_NETO'],
                        'factura': row['FACTURA_NORMALIZADA'],
                        'consumido': False,
                        'remaining_qty': row['CANT_ABS']
                    })
                
                idx_pedido_ean, idx_prov_ean, idx_ean_pedido, idx_ean_prov, idx_ean_solo = {}, {}, {}, {}, {}
                for linea in lineas_detalle:
                    key_1 = (linea['pedido'], linea['ean'])
                    if key_1 not in idx_pedido_ean: idx_pedido_ean[key_1] = []
                    idx_pedido_ean[key_1].append(linea)
                    
                    key_2 = (linea['proveedor'], linea['ean'])
                    if key_2 not in idx_prov_ean: idx_prov_ean[key_2] = []
                    idx_prov_ean[key_2].append(linea)
                    
                    key_3a = (linea['ean'], linea['pedido'])
                    if key_3a not in idx_ean_pedido: idx_ean_pedido[key_3a] = []
                    idx_ean_pedido[key_3a].append(linea)
                    
                    key_3b = (linea['ean'], linea['proveedor'])
                    if key_3b not in idx_ean_prov: idx_ean_prov[key_3b] = []
                    idx_ean_prov[key_3b].append(linea)
                    
                    key_3c = linea['ean']
                    if key_3c not in idx_ean_solo: idx_ean_solo[key_3c] = []
                    idx_ean_solo[key_3c].append(linea)

                pedidos_con_provision = {}
                for _, row in df_conta_orig.iterrows():
                    glosa_con = str(row['GLOSA']).upper()
                    doc_con = str(row['NUMDOCSUST']).upper()
                    match_p = re.search(r'(P\d+)', glosa_con)
                    if match_p and ("CON" in glosa_con or "CON" in doc_con):
                        num_con = "".join(re.findall(r'\d+', glosa_con if "CON" in glosa_con else doc_con))
                        pedidos_con_provision[match_p.group(1)] = f"CON-{num_con}"

                # PASO 5: CORRECCIÓN DEL KARDEX
                df_kardex_proc = df_kardex_orig.copy()
                c_u_corr, c_t_corr, f_corr, log_cambios = [], [], [], []
                
                for _, fila in df_kardex_proc.iterrows():
                    cod_tgu = str(fila['CodTGu']).strip()
                    ped = str(fila['PO']).strip()
                    prod_ean = extraer_ean(fila['CodDPr'])
                    prov_kardex = str(fila['REFERENCIA']).strip().upper()
                    
                    try: cant_kardex = float(str(fila['CntIte']).replace(',', '.')) if pd.notnull(fila['CntIte']) else 0.0
                    except: cant_kardex = 0.0
                    cant_kardex_abs = round(abs(cant_kardex), 4)
                    
                    try: costo_orig = float(str(fila['CoUMNc']).replace(',', '.')) if pd.notnull(fila['CoUMNc']) else 0.0
                    except: costo_orig = 0.0
                    
                    cambio, fact_exacta, costo_u = [], None, None
                
                    if cod_tgu in ['IC', 'SDP']:
                        # FILTRO 1
                        key_1 = (ped, prod_ean)
                        if key_1 in idx_pedido_ean:
                            for linea in idx_pedido_ean[key_1]:
                                if not linea['consumido'] and abs(linea['cant'] - cant_kardex_abs) < 0.005:
                                    linea['consumido'] = True
                                    linea['remaining_qty'] = 0.0
                                    fact_exacta, costo_u = linea['factura'], linea['costo']
                                    cambio.append("MATCH_FILTRO_1_PEDIDO")
                                    break
                        
                        # FILTRO 2
                        if costo_u is None:
                            key_2 = (prov_kardex, prod_ean)
                            if key_2 in idx_prov_ean:
                                for linea in idx_prov_ean[key_2]:
                                    if not linea['consumido'] and abs(linea['cant'] - cant_kardex_abs) < 0.005:
                                        linea['consumido'] = True
                                        linea['remaining_qty'] = 0.0
                                        fact_exacta, costo_u = linea['factura'], linea['costo']
                                        cambio.append("MATCH_FILTRO_2_PROVEEDOR")
                                        break
                        
                        # FILTRO 3
                        if costo_u is None:
                            linea_f3 = None
                            key_3a = (prod_ean, ped)
                            if key_3a in idx_ean_pedido:
                                for linea in idx_ean_pedido[key_3a]:
                                    if not linea['consumido'] and linea['remaining_qty'] > 0.005:
                                        linea_f3 = linea
                                        break
                            if linea_f3 is None:
                                key_3b = (prod_ean, prov_kardex)
                                if key_3b in idx_ean_prov:
                                    for linea in idx_ean_prov[key_3b]:
                                        if not linea['consumido'] and linea['remaining_qty'] > 0.005:
                                            linea_f3 = linea
                                            break
                            if linea_f3 is None:
                                if prod_ean in idx_ean_solo:
                                    for linea in idx_ean_solo[prod_ean]:
                                        if not linea['consumido'] and linea['remaining_qty'] > 0.005:
                                            linea_f3 = linea
                                            break
                            
                            if linea_f3 is not None:
                                linea_f3['remaining_qty'] = round(linea_f3['remaining_qty'] - cant_kardex_abs, 4)
                                if linea_f3['remaining_qty'] <= 0.005:
                                    linea_f3['consumido'] = True
                                    linea_f3['remaining_qty'] = 0.0
                                fact_exacta, costo_u = linea_f3['factura'], linea_f3['costo']
                                cambio.append("MATCH_FILTRO_3_EAN")
                
                        if costo_u is not None:
                            fact_f = fact_exacta
                            if abs(costo_u - costo_orig) > 0.01:
                                cambio.append("COSTO_CORREGIDO")
                        else:
                            costo_u = costo_orig
                            if ped in pedidos_con_provision:
                                fact_f = pedidos_con_provision[ped]
                                cambio.append("PROVISION")
                            else:
                                fact_f = normalizar_comprobante_dinamico(str(fila['FACT_REF']), pedido_ref=ped, es_sdp=(cod_tgu == 'SDP'), dict_prov=dict_pedido_a_proveedor, dict_pref=dict_prefijos) or "SIN_DOC"
                        
                        costo_t = cant_kardex * costo_u
                    else:
                        costo_u = costo_orig
                        try: costo_t = float(str(fila['CoTMNc']).replace(',', '.')) if pd.notnull(fila['CoTMNc']) else 0.0
                        except: costo_t = 0.0
                        fact_f = preservar_comprobante(str(fila['FACT_REF']))
                
                    c_u_corr.append(costo_u)
                    c_t_corr.append(costo_t)
                    f_corr.append(fact_f)
                    log_cambios.append(", ".join(cambio) if cambio else "SIN_CAMBIOS")
                
                df_kardex_proc['CoUMNc_CORRECTO'] = c_u_corr
                df_kardex_proc['CoTMNc_CORRECTO'] = c_t_corr
                df_kardex_proc['FACT_REF_FINAL']  = f_corr
                df_kardex_proc['CAMBIOS_AUDIT']   = log_cambios

                # PASO 6: EXPORTACIÓN EN MEMORIA Y ESTILOS OPENPYXL
                output_buffer = io.BytesIO()
                
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    df_kardex_proc.to_excel(writer, sheet_name='LOGISTICA', index=False)
                    df_detalle.to_excel(writer, sheet_name='DETALLE_FACTURAS_PROCESADO', index=False)
                    df_conta_orig.to_excel(writer, sheet_name='ORIGINAL_CONTABILIDAD', index=False)
                
                output_buffer.seek(0)
                wb = load_workbook(output_buffer)
                
                # Estilos
                font_blanca_bold = Font(name="Segoe UI", size=11, color="FFFFFF", bold=True)
                font_regular = Font(name="Segoe UI", size=10)
                fill_negro = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
                fill_amarillo_titulo = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                fill_azul_titulo = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                fill_verde_f1 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                fill_azul_f2 = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                fill_naranja_f3 = PatternFill(start_color="FFE4B0", end_color="FFE4B0", fill_type="solid")
                border_fino = Border(
                    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),  bottom=Side(style='thin', color='D9D9D9')
                )

                # Aplicar estilos a DETALLE_FACTURAS_PROCESADO
                ws_det = wb['DETALLE_FACTURAS_PROCESADO']
                headers_det = [c.value for c in ws_det[1]]
                ws_det.row_dimensions[1].height = 26
                columnas_nuevas_det = ['PEDIDO_CORRECTO', 'EAN_LIMPIO', 'COSTO_UNITARIO_NETO', 'FACTURA_NORMALIZADA', 'CANT_ABS', 'PROV_LIMPIO']
                for col_idx in range(1, len(headers_det) + 1):
                    cell = ws_det.cell(row=1, column=col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if headers_det[col_idx - 1] in columnas_nuevas_det:
                        cell.font, cell.fill = font_blanca_bold, fill_azul_titulo
                    else:
                        cell.font, cell.fill = font_blanca_bold, fill_negro
                idx_costo_u_det = headers_det.index('COSTO_UNITARIO_NETO') + 1
                for r in range(2, ws_det.max_row + 1):
                    ws_det.row_dimensions[r].height = 19
                    ws_det.cell(r, idx_costo_u_det).number_format = '"S/." #,##0.0000'
                    for c in range(1, len(headers_det) + 1):
                        ws_det.cell(r, c).font = font_regular
                        ws_det.cell(r, c).border = border_fino
                for col in ws_det.columns:
                    ws_det.column_dimensions[col[0].column_letter].width = 22

                # Aplicar estilos a LOGISTICA
                ws_kardex = wb['LOGISTICA']
                headers_kardex = [c.value for c in ws_kardex[1]]
                columnas_nuevas_k = ['CoUMNc_CORRECTO', 'CoTMNc_CORRECTO', 'FACT_REF_FINAL', 'CAMBIOS_AUDIT']
                ws_kardex.row_dimensions[1].height = 26
                for col_idx in range(1, len(headers_kardex) + 1):
                    cell = ws_kardex.cell(row=1, column=col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if headers_kardex[col_idx - 1] in columnas_nuevas_k:
                        cell.font, cell.fill = font_blanca_bold, fill_amarillo_titulo
                    else:
                        cell.font, cell.fill = font_blanca_bold, fill_negro
                idx_c_u_k = headers_kardex.index('CoUMNc_CORRECTO') + 1
                idx_c_t_k = headers_kardex.index('CoTMNc_CORRECTO') + 1
                idx_fact_k = headers_kardex.index('FACT_REF_FINAL') + 1
                idx_audit_k = headers_kardex.index('CAMBIOS_AUDIT') + 1

                for r in range(2, ws_kardex.max_row + 1):
                    ws_kardex.row_dimensions[r].height = 19
                    ws_kardex.cell(r, idx_c_u_k).number_format = '"S/." #,##0.0000'
                    ws_kardex.cell(r, idx_c_t_k).number_format = '"S/." #,##0.00'
                    for c in range(1, len(headers_kardex) + 1):
                        ws_kardex.cell(r, c).border = border_fino
                        ws_kardex.cell(r, c).font = font_regular
                        
                    audit_val = str(ws_kardex.cell(r, idx_audit_k).value)
                    if "MATCH_FILTRO_1_PEDIDO" in audit_val:
                        ws_kardex.cell(r, idx_fact_k).fill = fill_verde_f1
                    elif "MATCH_FILTRO_2_PROVEEDOR" in audit_val:
                        ws_kardex.cell(r, idx_fact_k).fill = fill_azul_f2
                    elif "MATCH_FILTRO_3_EAN" in audit_val:
                        ws_kardex.cell(r, idx_fact_k).fill = fill_naranja_f3
                        
                    if "COSTO_CORREGIDO" in audit_val:
                        ws_kardex.cell(r, idx_c_u_k).fill = fill_verde_f1
                for col in ws_kardex.columns:
                    ws_kardex.column_dimensions[col[0].column_letter].width = 22

                # Guardar resultado final en buffer para descarga
                final_buffer = io.BytesIO()
                wb.save(final_buffer)
                final_buffer.seek(0)

                st.success("¡Conciliación finalizada con éxito!")
                
                st.download_button(
                    label="📥 Descargar Excel Conciliado",
                    data=final_buffer,
                    file_name="Reporte_Kardex_Conciliado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"Ocurrió un error durante el procesamiento: {e}")